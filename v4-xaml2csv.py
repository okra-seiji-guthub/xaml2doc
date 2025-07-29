## v3-xaml2csv.py with TryCatch support
import xml.etree.ElementTree as ET
import sys
import os
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
import copy


Gtag_name = ""

EXCLUDED_PROPERTY_PREFIXES = (
    '{http://schemas.microsoft.com/netfx/2009/xaml/activities/presentation}',
    '{http://schemas.microsoft.com/netfx/2010/xaml/activities/presentation}',
    '{http://schemas.microsoft.com/winfx/2006/xaml}',
    '{http://schemas.microsoft.com/netfx/2010/xaml/activities/debugger}'
)

NS = {
    'x': 'http://schemas.microsoft.com/winfx/2006/xaml',
    'scg': "clr-namespace:System.Collections.Generic;assembly=mscorlib",
    'ui': 'http://schemas.uipath.com/workflow/activities',
    'xamlp': 'http://schemas.microsoft.com/netfx/2009/xaml/activities',
}

SPECIAL_FLAT_TAGS = {"Flowchart", "FlowDecision",  "ExcelApplicationScope",
                     "Sequence", "FlowStep", "TryCatch","TryCatch.Try","TryCatch.Catches",
                     "FlowDecision.True", "FlowDecision.False",
                     "If","If.Then", "If.Else", "If.Condition",
                     "ForEach","ForEachRow","ForEach.Body",
                     "InvokeWorkflowFile"}


#EXCLUDED_PATH_KEYWORDS = {"Flowchart.Variables", "WorkflowViewStateService", "Flowchart.StartNode"}
EXCLUDED_PATH_KEYWORDS = {"Flowchart.Variables", "WorkflowViewStateService"}


VARIABLE_TAG = 'Variable'

def extract_arguments(root):
    arguments = []
    members = root.find('x:Members', namespaces=NS)
    if members is not None:
        for prop in members.findall('x:Property', namespaces=NS):
            name = prop.attrib.get('Name')
            typ = prop.attrib.get('Type')
            if name and typ:
                arguments.append((name, typ))
    return arguments

# def extract_variables(elem, current_scope, variables):
#     for child in elem:
#         tag_short = child.tag.split('}')[-1]
#         if tag_short == VARIABLE_TAG:
#             print(f"## Variable found: {child.attrib.get('Name')}:{child.attrib.get(f'{{NS['x']}}TypeArguments')} in scope: {current_scope}")
#             name = child.attrib.get('Name')
#             typ = child.attrib.get(f'{{NS['x']}}TypeArguments')
#             if name and typ:
#                 print(f"######## Variable: {name}, Type: {typ}, Scope: {current_scope}")
#                 variables.append((name, typ, current_scope))
#         extract_variables(child, current_scope, variables)


def parse_table_xml(xml_string):
    # 名前空間を定義（xmlns定義されているので）
    namespaces = {
        'xs': 'http://www.w3.org/2001/XMLSchema',
        'msdata': 'urn:schemas-microsoft-com:xml-msdata'
    }

    translated_string = ""

    root = ET.fromstring(xml_string)

    # --- 列情報の抽出 ---
    schema = root.find('xs:schema', namespaces)
    columns = []

    if schema is not None:
        for element in schema.findall('.//xs:element[@name="TableName"]/xs:complexType/xs:sequence/xs:element', namespaces):
            column_name = element.attrib.get('name')
            column_type = element.attrib.get('type', 'xs:string')  # デフォルトでstring
            columns.append({
                'name': column_name,
                'type': column_type.replace('xs:', '')
            })

    print("列情報:")
    translated_string += "列情報:["
    for col in columns:
        translated_string += f"{col['name']} ({col['type']} ), "
        print(f" - {col['name']} ({col['type']})")
    translated_string += "]<NL/>"

    # --- データの抽出 ---
    print("\nデータ:")
    translated_string += "データ:["
    for table_elem in root.findall('TableName'):
        row = {}
        for col in columns:
            cell = table_elem.find(col['name'])
            row[col['name']] = cell.text if cell is not None else None
        translated_string += f"{row}, "
        print(row)
    translated_string += "]"
    return translated_string


def parse_invoke_args_xml(root):
    # pase <ui:InvokeWorkflowFile.Arguments>
    columns = []

    for element in root:
        # text = element.text.strip() if element.text else ''
        arg_text = element.text
        arg_type = element.attrib.get(f"{{{NS['x']}}}TypeArguments")
        arg_key  = element.attrib.get(f"{{{NS['x']}}}Key")
        arg_direction = ''
        if element.tag.endswith('InArgument'):
            arg_direction = 'In'
        else:
            arg_direction = 'Out'
        columns.append({
            'key': arg_key,
            'type': arg_type,
            'direction': arg_direction,
            'text': arg_text})

    str_props = "<NL/>".join([f" {col['key']} = {col['text']} / {col['type']}:{col['direction']}"  for col in columns])
    
    return str_props

def flatten_child_properties(elem, skip=False):
    if skip or elem is None:
        return {}
    props = {}
    try:
        for child in elem:
            tag = child.tag.split('}')[-1]
            if tag.startswith(tuple(EXCLUDED_PATH_KEYWORDS)):
                continue
            text = (child.text or '').strip()
            if len(child):
                nested = flatten_child_properties(child)
                for k, v in nested.items():
                    props[f"{tag}.{k}"] = v
            elif text:
                props[tag] = text
     
    except Exception as e:
        print(f"####Error Activity:{Gtag_name} / parsing element {elem}: {e}")
        return {}   
    return props

def translate(text):
    EGNORE_PREFIXES = ["x:", "scg"]
    TRANSLATE_WORDS = {
    "new ": "初期化した ", 
    "vbCrLf":"改行文字",
    "Null":"",
    "CLICK_SINGLE": "シングルクリック",
    "CLICK_DOUBLE": "ダブルクリック",
    "CLICK_DOWN": "マウスダウン",
    "CLICK_UP": "マウスアップ",
    "BTN_LEFT": "左ボタン",
    "BTN_MIDDLE": "中央ボタン",
    "BTN_RIGHT": "右ボタン",
    "x:":""}

    for prefix in EGNORE_PREFIXES:
        if text.startswith(prefix):
            return text.replace(prefix, "")

    for key, value in TRANSLATE_WORDS.items():
        text = text.replace(key, value)
    return text

def get_props_from_dict(props, dict):
    for k, v in props.items():
        if k in dict.keys():
            dict[k] = translate(v)
    return dict

def prop(format_str, prop):
    return format_str.format(prop) if prop else ""

def get_Assign_props(props):
    DICT ={"Assign.To.OutArgument":"",
           "Assign.Value.InArgument":""}
    dict = get_props_from_dict(props, DICT)
    
    return f"値の設定をおこなう：設定先{dict['Assign.To.OutArgument']}<NL/>設定元[{dict['Assign.Value.InArgument']}]"

def get_If_props(props):
    DICT ={"Condition":""}
    dict = get_props_from_dict(props, DICT)
    
    return f"分岐条件：{dict['Condition']} の場合に分岐をおこなう"

def get_AppendLine_props(props):
    DICT ={"Encoding":"",
           "FileName":"",
           "Text":""}
    dict = get_props_from_dict(props, DICT)
    
    return (f"ファイル書き込み：ファイル名{dict['FileName']}に<NL/>"
            f"{dict['Text']}を{dict['Encoding']}で追加書き込み")

def get_PathExists_props(props):
    DICT = {"Exists":"",
            "Path":"",
            "PathType":""}
    dict = get_props_from_dict(props, DICT)
    
    return (f"ファイルの存在確認：パス{dict['Path']}の存在を確認し<NL/>"
            f"{dict['Exists']} 変数にTrueを返す / パスタイプ：{dict['PathType']}")

def get_ForEach_props(props):
    DICT = {"Values":""}
    dict = get_props_from_dict(props, DICT)
    
    return f"全件繰り返し：全件処理するデータ：{dict['Values']}"

def get_FlowDecision_props(props):
    DICT ={"Condition":""}
    dict = get_props_from_dict(props, DICT)
    
    return f"分岐条件：{dict['Condition']} の場合に分岐をおこなう"

def get_ExcelApplicationScope_props(props):
    DICT = {"Password":"",
            "AutoSave":"",
            "CreateNewFile":"",
            "ReadOnly":"",
            "Workbook":"",
            "WorkbookPath":""}
    dict = get_props_from_dict(props, DICT)
    
    return (f"Excelでファイルを操作: ファイルパス{dict['WorkbookPath']}を<NL/>"
            f"{dict['ReadOnly']}で開き、{dict['AutoSave']}で保存する / パスワード：{dict['Password']}"
            f" / 新規作成：{dict['CreateNewFile']} / ワークブック名：{dict['Workbook']}")


def get_ReadRange_props(props):
    DICT = {"Range":"",
            "AddHeaders":"",
            "DataTable":"",
            "SheetName":"",
            "WorkbookPath":""}
    dict = get_props_from_dict(props, DICT)
    
    return (f"Excelの範囲を読み込み: ワークブックパス{dict['WorkbookPath']} / シート名{dict['SheetName']}の範囲[{dict['Range']}]を<NL/>"
            f"データテーブル{dict['DataTable']}に読み込み / ヘッダー行あり：{dict['AddHeaders']}")

def get_ExcelReadRange_props(props):
    DICT = {"AddHeaders":"",
            "DataTable":"",
            "Range":"",
            "SheetName":""}
    dict = get_props_from_dict(props, DICT)
    
    return (f"Excelの範囲を読み込み: シート名{dict['SheetName']}の範囲[{dict['Range']}]を"
            f"データテーブル{dict['DataTable']}に読み込み / ヘッダー行あり：{dict['AddHeaders']}")

def get_ExcelWriteRange_props(props):
    DICT = {"AddHeaders":"",
            "DataTable":"",
            "SheetName":"",
            "StartingCell":""}
    dict = get_props_from_dict(props, DICT)
    
    return (f"Excelの範囲に書き込み: シート名{dict['SheetName']}の開始セル[{dict['StartingCell']}]に"
            f"データテーブル{dict['DataTable']}を書き込み / ヘッダー行あり：{dict['AddHeaders']}")

def get_ExcelDeleteRange_props(props):
    DICT = {"Range":"",
            "SheetName":"",
            "ShiftCells":"",
            "ShiftOption":""}
    dict = get_props_from_dict(props, DICT)
    
    return (f"Excelの範囲の削除: シート名{dict['SheetName']}の範囲[{dict['Range']}]を削除し、"
            f"セルを{dict['ShiftCells']}方向に{dict['ShiftOption']}する")

def get_ExcelSelectRange_props(props):
    DICT = {"Range":"",
            "SheetName":""}
    dict = get_props_from_dict(props, DICT)
    
    return f"Excelの範囲を選択: シート名{dict['SheetName']}の範囲[{dict['Range']}]を選択する"

def get_ExcelReadCell_props(props):
    DICT = {"Cell":"",
            "SheetName":"",
            "ExcelReadCell.Result.OutArgument":""}
    dict = get_props_from_dict(props, DICT)
    
    return (f"Excelのセルを読み込み: シート名{dict['SheetName']}のセル[{dict['Cell']}]を読み込み、"
            f"結果を変数{dict['ExcelReadCell.Result.OutArgument']}に格納する")

def get_ExcelWriteCell_props(props):
    DICT = {"Cell":"",
            "SheetName":"",
            "Text":""}
    dict = get_props_from_dict(props, DICT)
    
    return (f"Excelのセルに書き込み: シート名{dict['SheetName']}のセル[{dict['Cell']}]に、テキスト: <NL/>"
            f"[{dict['Text']}]を書き込む")

def get_ExcelCloseWorkbook_props(props):
    DICT = {"Workbook":""}
    dict = get_props_from_dict(props, DICT)
    
    return f"Excelのワークブックを閉じる: ワークブック{dict['Workbook']}を閉じる"

def get_ExcelSaveWorkbook_props(props):
    DICT = {}
    dict = get_props_from_dict(props, DICT)
    
    return f"Excelのワークブックを保存: 現在のワークブックを保存する"

def get_OpenApplication_props(props):
    DICT = {"ApplicationWindow":"",
            "Arguments":"",
            "TimeoutMS":"",
            "WorkingDirectory":"",
            "FileName":"",
            "Selector":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"アプリケーションを起動：ファイル名{dict['FileName']}を起動 / オプション：引数[{dict['Arguments']}]<NL/>"
            f" ワーキングディレクトリ: {dict['WorkingDirectory']} / タイムアウト(ms): {dict['TimeoutMS']}<NL/>"
            f" アプリケーションウィンドウ: {dict['ApplicationWindow']} セレクター: {dict['Selector']}")

def get_WindowScope_props(props):
    DICT = {"SearchScope":"",
            "TimeoutMS":"",
            "Window":"",
            "ApplicationWindow":"",
            "Selector":"",
            "SendWindowMessages":"",
            "SimulateClick":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"ウィンドウスコープ: アプリウインドウ: {dict['ApplicationWindow']} / ウインドウ: [{dict['Window']}]を操作する<NL/>"
            f" スコープ: {dict['SearchScope']} / タイムアウト(ms): {dict['TimeoutMS']} / セレクター: {dict['Selector']}")

def get_CloseApplication_props(props):
    DICT = {}
    
    dict = get_props_from_dict(props, DICT)

    return f"アプリケーションを閉じる"

def get_TypeInto_props(props):
    DICT = {"DelayBefore":"",
            "DelayBetweenKeys":"",
            "DelayMS":"",
            "Activate":"",
            "ClickBeforeTyping":"",
            "EmptyField":"",
            "SendWindowMessages":"",
            "SimulateType":"",
            "Text":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"テキスト入力: テキスト[{dict['Text']}]を入力する<NL/>"
            f" 遅延(ms): {dict['DelayMS']} / 遅延前: {dict['DelayBefore']} / 遅延キー間: {dict['DelayBetweenKeys']}"
            f" / アクティブ化: {dict['Activate']} / クリック前入力: {dict['ClickBeforeTyping']} / フィールド空にする: {dict['EmptyField']}"
            f" / ウィンドウメッセージ送信: {dict['SendWindowMessages']} / シミュレート入力: {dict['SimulateType']}")

def get_Click_props(props):
    DICT = {"DelayBefore":"",
            "DelayMS":"",
            "ClickType":"",
            "KeyModifiers":"",
            "MouseButton":"",
            "SendWindowMessages":"",
            "SimulateClick":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"クリック操作: クリックタイプ[{dict['ClickType']}]、マウスボタン: {dict['MouseButton']} でクリックする<NL/>"
            f" 遅延(ms): {dict['DelayMS']} / 遅延前: {dict['DelayBefore']} / キーモディファイア: {dict['KeyModifiers']}"
            f" / ウィンドウメッセージ送信: {dict['SendWindowMessages']} / シミュレートクリック: {dict['SimulateClick']}")

def get_SendHotkey_props(props):
    DICT = {"DelayBefore":"",
            "DelayBetweenKeys":"",
            "DelayMS":"",
            "Activate":"",
            "ClickBeforeTyping":"",
            "EmptyField":"",
            "Key":"",
            "KeyModifiers":"",
            "SpecialKey":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"ホットキー送信: キーモディファイア: {dict['KeyModifiers']}, ホットキー[{dict['Key']}]を<NL/>"
            f" 特殊キー: {dict['SpecialKey']} で送信する<NL/>"
            f" 遅延(ms): {dict['DelayMS']} / 遅延前: {dict['DelayBefore']}"
            f" / 遅延キー間: {dict['DelayBetweenKeys']} / アクティブ化: {dict['Activate']}"
            f" / クリック前入力: {dict['ClickBeforeTyping']} / フィールド空にする: {dict['EmptyField']}"
            f" / キーモディファイア: {dict['KeyModifiers']}")

def get_BuildDataTable_props(props):
    DICT = {"DataTable":"",
            "TableInfo":""}
    
    dict = get_props_from_dict(props, DICT)
    table_info = parse_table_xml(dict['TableInfo']) if dict['TableInfo'] else "なし"

    return (f"テーブル作成作成: {dict['DataTable']}を構築"
            f" テーブル情報:<NL/>{table_info}")


def get_MergeDataTable_props(props):
    DICT = {"Destination":"",
            "MissingSchemaAction":"",
            "Source":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"データテーブルのマージ: ソース{dict['Source']}を宛先{dict['Destination']}にマージする<NL/>"
            f" マージ方法: {dict['MissingSchemaAction']}")

def get_ClearDataTable_props(props):
    DICT = {"DataTable":""}
    
    dict = get_props_from_dict(props, DICT)

    return f"データテーブルのクリア: データテーブル{dict['DataTable']}をクリアする"

def get_InvokeWorkflowFile_props(props):
    DICT = {"ContinueOnError":"",
            "WorkflowFileName":"",
            "InvokeWorkflowFile.Arguments":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"ワークフロー呼び出し: ファイル：{dict['WorkflowFileName']} / エラー継続: {dict['ContinueOnError']}"
            f" / 起動引数:<NL/>{dict['InvokeWorkflowFile.Arguments']}")

def get_TerminateWorkflow_props(props):
    DICT = {"Reason":""}
    
    dict = get_props_from_dict(props, DICT)

    return f"ワークフローの終了: 理由: {dict['Reason']}"

def get_MessageBox_props(props):
    DICT = {"Caption":"",
            "ChosenButton":"",
            "Buttons":"",
            "Text":"",
            "TopMost":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"メッセージボックスを表示:<NL/>"
            f" メッセージ: {dict['Text']}<NL/>"
            f" ボタン: {dict['Buttons']} / キャプション: {dict['Caption']} / 選択ボタン: {dict['ChosenButton']} "
            f"/ TopMost: {dict['TopMost']}")

def get_CreateFile_props(props):
    DICT = {"ContinueOnError":"",
            "Name":"",
            "Path":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"ファイルの作成: パス{dict['Path']}にファイル{dict['Name']}を作成する<NL/>"
            f" 継続エラー: {dict['ContinueOnError']}")

def get_CopyFile_props(props):
    DICT = {"ContinueOnError":"",
            "Destination":"",
            "Overwrite":"",
            "Path":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"ファイルのコピー: コピー元パス:<NL/>"
            f" {dict['Path']}<NL/>"
            f"コピー先パス:<NL/>"
            f" {dict['Destination']}にコピーする<NL/>"
            f" 継続エラー: {dict['ContinueOnError']} / 上書き: {dict['Overwrite']}")

def get_MoveFile_props(props):
    DICT = {"ContinueOnError":"",
            "Destination":"",
            "Overwrite":"",
            "Path":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"ファイルの移動: 移動元パス:<NL/>"
            f" {dict['Path']}<NL/>"
            f"移動先パス:<NL/>"
            f" {dict['Destination']}<NL/>"
            f" 継続エラー: {dict['ContinueOnError']} / 上書き: {dict['Overwrite']}")

def get_CreateDirectory_props(props):
    DICT = {"ContinueOnError":"",
            "Path":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"ディレクトリの作成: パス{dict['Path']}にディレクトリを作成する<NL/>"
            f" 継続エラー: {dict['ContinueOnError']}")

def get_WaitUiElementAppear_props(props):
    DICT = {"FoundElement":"",
            "WaitActive":"",
            "WaitVisible":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"UI要素の出現待ち: 要素{dict['FoundElement']}が出現するまで待機<NL/>" 
            f" アクティブ化: {dict['WaitActive']} / 可視化: {dict['WaitVisible']}")

def get_SetToClipboard_props(props):
    DICT = {"Text":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"クリップボードに設定: テキスト {dict['Text']} をクリップボードに設定する")

def get_AddToCollection_props(props):
    DICT = {"Collection":"",
            "Item":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"コレクションに追加: アイテム: {dict['Item']} を、コレクション: {dict['Collection']} に追加する")


def get_ClearCollection_props(props):
    DICT = {"Collection":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"コレクションのクリア: コレクション{dict['Collection']}をクリアする")

def get_Check_props(props):
    DICT = {"ContinueOnError":"",
            "DelayAfter":"",
            "DelayBefore":"",
            "Action":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"チェック操作: アクション{dict['Action']}を実行する<NL/>"
            f" 継続エラー: {dict['ContinueOnError']} / 遅延後: {dict['DelayAfter']} / 遅延前: {dict['DelayBefore']}")

def get_UiElementExists_props(props):
    DICT = {"Exists":"",
            "ClippingRegion":"",
            "Element":"",
            "Selector":"",
            "TimeoutMS":"",
            "WaitForReady":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"UI要素の存在確認: セレクター: {dict['Selector']} の存在を確認し<NL/>"
            f" 結果を変数{dict['Exists']}に格納する / 要素: {dict['Selector']} / "
            f" タイムアウト(ms): {dict['TimeoutMS']} / クリッピング領域: {dict['ClippingRegion']} / "
            f" 準備完了待ち: {dict['WaitForReady']}")

def get_SelectItem_props(props):
    DICT = {"ContinueOnError":"",
            "DelayAfter":"",
            "DelayBefore":"",
            "Item":"",
            "SelectItem.Items.List.String":"",
            "Items":""}
    
    dict = get_props_from_dict(props, DICT)

    return (f"アイテム選択: アイテム{dict['Item']} を選択する<NL/>"
            f" アイテムリスト: {dict['Items']}<NL/>"
            f" 継続エラー: {dict['ContinueOnError']} / 遅延後: {dict['DelayAfter']} / 遅延前: {dict['DelayBefore']}")


def get_props(tag_name, combined_props):
    GENERATIVE_TAGS = {"Assign":get_Assign_props,
                       "If":get_If_props,
                       "AppendLine":get_AppendLine_props,
                       "PathExists":get_PathExists_props,
                       "ForEach":get_ForEach_props,
                       "FlowDecision":get_FlowDecision_props,
                       "ExcelApplicationScope":get_ExcelApplicationScope_props,
                       "ReadRange":get_ReadRange_props,
                       "ExcelReadRange":get_ExcelReadRange_props,
                       "ExcelWriteRange":get_ExcelWriteRange_props,
                       "ExcelDeleteRange":get_ExcelDeleteRange_props,
                       "ExcelSelectRange":get_ExcelSelectRange_props,
                       "ExcelReadCell":get_ExcelReadCell_props,
                       "ExcelWriteCell":get_ExcelWriteCell_props,
                       "ExcelCloseWorkbook":get_ExcelCloseWorkbook_props,
                       "ExcelSaveWorkbook":get_ExcelSaveWorkbook_props,
                       "OpenApplication":get_OpenApplication_props,
                       "WindowScope":get_WindowScope_props,
                       "CloseApplication":get_CloseApplication_props,
                       "TypeInto":get_TypeInto_props,
                       "Click":get_Click_props,
                       "SendHotkey":get_SendHotkey_props,
                       "BuildDataTable":get_BuildDataTable_props,
                       "MergeDataTable":get_MergeDataTable_props,
                       "ClearDataTable":get_ClearDataTable_props,
                       "InvokeWorkflowFile":get_InvokeWorkflowFile_props,
                       "TerminateWorkflow":get_TerminateWorkflow_props,
                       "MessageBox":get_MessageBox_props,
                       "CopyFile":get_CopyFile_props,
                       "MoveFile":get_MoveFile_props,
                       "CreateDirectory":get_CreateDirectory_props,
                       "WaitUiElementAppear":get_WaitUiElementAppear_props,
                       "SetToClipboard":get_SetToClipboard_props,
                       "AddToCollection":get_AddToCollection_props,
                       "ClearCollection":get_ClearCollection_props,
                       "Check":get_Check_props,
                       "UiElementExists":get_UiElementExists_props,
                       "SelectItem":get_SelectItem_props,
                       "CreateFile":get_CreateFile_props}
    
    str_props = "" 
    if tag_name in GENERATIVE_TAGS:
        str_props = GENERATIVE_TAGS[tag_name](combined_props) 
    else:
        str_props = ', '.join([f"{k}={translate(v)}" for k, v in combined_props.items()])

    return str_props


def collect_activity_details(elem, path_stack, index, rows, variables, parent_type=None):
    display_name = elem.attrib.get('DisplayName')
    path_name = display_name
    tag_short = elem.tag.split('}')[-1]

    if any(excl in tag_short for excl in EXCLUDED_PATH_KEYWORDS):
        return

    if  tag_short.endswith(('FlowDecision.True', 'FlowDecision.False', 'TryCatch.Try', 'TryCatch.Catches')):
        display_name = tag_short
        path_name = tag_short.split('.')[1]

    if tag_short.startswith('If'):
        display_name = tag_short
        path_name = display_name
    elif tag_short.startswith('ForEach'):
        display_name = tag_short
        path_name = display_name
        #print(f"## ForEachRow found: {display_name} in path: {path_stack}"  )

    if display_name:
        path = '>'.join(path_stack)
        attrib_props = {
            k: v for k, v in elem.attrib.items()
            if k != 'DisplayName' and not k.startswith(EXCLUDED_PROPERTY_PREFIXES)
        }
        if tag_short == "UiElementExists":
            elem = elem.find(".//ui:UiElementExists.Target", namespaces=NS)
            elem = elem.find(".//ui:Target", namespaces=NS)
            for k, v in elem.attrib.items():
                if k.startswith(EXCLUDED_PROPERTY_PREFIXES):
                    continue
                attrib_props[k] = v

        if tag_short == "SelectItem":
            elem = elem.find(".//ui:SelectItem.Items", namespaces=NS)
            elem = elem.find(".//scg:List", namespaces=NS)
            if elem is not None:
                attrib_props['Items'] = f"[{', '.join([item.text for item in elem if item.text])}]"

        if tag_short == "InvokeWorkflowFile":
            elem = elem.find(".//ui:InvokeWorkflowFile.Arguments", namespaces=NS)
            if elem is not None:
                attrib_props['InvokeWorkflowFile.Arguments'] = parse_invoke_args_xml(elem)
            
        skip_nested = tag_short in SPECIAL_FLAT_TAGS
        Gtag_name = display_name
        child_props = flatten_child_properties(elem, skip=skip_nested)
        combined_props = {**attrib_props, **child_props}
        #prop_str = ', '.join([f"{k}={v}" for k, v in combined_props.items()])
        prop_str = get_props(tag_short, combined_props)
        rows.append([display_name, path, tag_short, prop_str])
        index[0] += 1

        scope = path if path else display_name
        # extract_variables(elem, scope, variables)

        path_stack.append(path_name)

    # ExcelApplicationScope
    if tag_short == "ExcelApplicationScope":
        body = elem.find(".//ui:ExcelApplicationScope.Body", namespaces=NS)
        if body is not None:
            action = body.find(".//ui:ActivityAction", namespaces=NS)
            if action is not None:
                seq = action.find(".//ui:Sequence", namespaces=NS)
                if seq is not None:
                    collect_activity_details(seq, path_stack[:], index, rows, tag_short)

    if tag_short == "ForEach":
        elem = elem.find(".//ui:ForEach.Body", namespaces=NS)
        #if body is not None:
        #collect_activity_details(body, path_stack[:], index, rows, tag_short)
    if tag_short == "ForEachRow":
        elem= elem.find(".//ui:ForEachRow.Body", namespaces=NS)

    if elem is None:
        return
    
    for child in elem:
        collect_activity_details(child, path_stack[:], index, rows, tag_short)

    if display_name:
        path_stack.pop()

def parse_xaml_to_csv(xaml_file, csv_file):
    if not os.path.exists(xaml_file):
        print(f"Error: File '{xaml_file}' not found.")
        return

    tree = ET.parse(xaml_file)
    root = tree.getroot()
    index = [1]
    rows = []
    variables = []

    arguments = extract_arguments(root)

    for child in root.iter():
        if child.attrib.get('DisplayName'):
            collect_activity_details(child, [], index, rows, variables)
            break

    with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)

        if arguments:
            writer.writerow(['Arguments'])
            writer.writerow(['Name', 'Type'])
            for arg in arguments:
                writer.writerow(arg)
            writer.writerow([])

        if variables:
            writer.writerow(['Variables'])
            writer.writerow(['Name', 'Type', 'Scope'])
            for var in variables:
                writer.writerow(var)
            writer.writerow([])

        #writer.writerow(['Path', 'DisplayName', 'Type', 'Properties'])
        writer.writerow(['DisplayName', 'Path', 'Type', 'Properties'])
        writer.writerows(rows)

    print(f"✅ CSV exported to: {csv_file}")

def write_csv_to_excel(csv_file, excel_template, output_excel):
    if not os.path.exists(excel_template):
        print(f"⚠️ Excel template not found: {excel_template}")
        return

    shutil.copyfile(excel_template, output_excel)
    wb = load_workbook(output_excel)
    ws = wb.active

    # 明示的にフォントを指定（MS ゴシック、サイズ10）
    forced_font = Font(name='メイリオ', size=8)
    with open(csv_file, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = forced_font  # ← 明示的にフォントを指定

    wb.save(output_excel)
    print(f"📄 Excel exported to: {output_excel}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python parse_xaml_to_csv.py <xaml filename>")

        xaml_file = "./AcMonthlyDataOutput/Main.xaml"
        base_name = os.path.splitext(xaml_file)[0]
        pathArray = base_name.split("/")
        module = pathArray.pop()
        senario = pathArray.pop()
        csv_file = f"./dest/{senario}-{module}.csv"
        parse_xaml_to_csv(xaml_file, csv_file)

        template_path = "./template/template.xlsx"
        output_excel = f"./dest/{senario}-{module}.xlsx"
        write_csv_to_excel(csv_file, template_path, output_excel)




    else:
        xaml_file = sys.argv[1]
        base_name = os.path.splitext(xaml_file)[0]
        pathArray = base_name.split("/")
        module = pathArray.pop()
        senario = pathArray.pop()
        csv_file = f"./dest/{senario}-{module}.csv"
        parse_xaml_to_csv(xaml_file, csv_file)

        template_path = "./template/template.xlsx"
        output_excel = f"./dest/{senario}-{module}.xlsx"
        write_csv_to_excel(csv_file, template_path, output_excel)
